"""
Örnek Kullanıcı Listesi Dosyaları
Her biçim desteklenmektedir
"""

# örnek.txt dosyası oluşturma Python kodu
example_users = [
    "john_doe",
    "maria_smith",
    "alex_jones",
    "user_123",
    "test_user",
    "telegram_user",
    "example_account"
]

# TXT formatında
with open('example_users.txt', 'w', encoding='utf-8') as f:
    for user in example_users:
        f.write(user + '\n')

print("✓ example_users.txt oluşturuldu")

# CSV formatında
import csv
with open('example_users.csv', 'w', newline='', encoding='utf-8') as f:
    writer = csv.writer(f)
    writer.writerow(['username'])  # Başlık
    for user in example_users:
        writer.writerow([user])

print("✓ example_users.csv oluşturuldu")

# XLSX formatında
try:
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = 'username'
    for i, user in enumerate(example_users, 2):
        ws[f'A{i}'] = user
    wb.save('example_users.xlsx')
    print("✓ example_users.xlsx oluşturuldu")
except ImportError:
    print("⚠ openpyxl yüklü değil, XLSX oluşturulmadı")
